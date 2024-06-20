import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the Excel file
file_path = 'Material Flow.xlsx'  # Update with your file path
df = pd.read_excel(file_path, sheet_name='For Alonzo', header=None)  # Read without headers

# Extract the relevant columns starting from the specified rows
parts = df.iloc[4:, 0].dropna().values  # Column A starting from row 5 (0-based index 4)
times = df.iloc[4:, 8].dropna().values  # Column I starting from row 5 (0-based index 8)

# Round the times to 2 decimal places
times = [round(time, 2) for time in times]

# Combine the parts and times into a list of tuples and sort by time in descending order
parts_times = sorted(list(zip(parts, times)), key=lambda x: x[1], reverse=True)

# Define the target time and the number of machines
target_time = 7.5
num_machines = 6

# Initialize machines
machines = [[] for _ in range(num_machines)]
machine_times = [0] * num_machines

# Assign parts to machines
for part, time in parts_times:
    # Find the machine with the current lowest total time that can accommodate this part
    for i in range(num_machines):
        if machine_times[i] + time <= target_time:
            machines[i].append((part, time))
            machine_times[i] += time
            break

# Prepare the results as a DataFrame
results = []
for i, machine in enumerate(machines):
    for part, time in machine:
        results.append([f"Machine {i + 1}", part, time])
results_df = pd.DataFrame(results, columns=["Machine", "Part", "Time"])

# Load the workbook and add the results to a new sheet
wb = load_workbook(file_path)
if "Results" in wb.sheetnames:
    del wb["Results"]  # Remove the existing "Results" sheet if it exists
ws = wb.create_sheet(title="Results")

# Write the DataFrame to the new sheet
for r in dataframe_to_rows(results_df, index=False, header=True):
    ws.append(r)

# Save the workbook
wb.save(file_path)

print("Results have been written to the 'Results' sheet.")
